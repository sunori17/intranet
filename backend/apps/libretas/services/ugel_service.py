# backend/apps/libretas/services/ugel_service.py
from __future__ import annotations
from io import BytesIO
from typing import List, Tuple
from openpyxl import load_workbook

from .excel_adapter import leer_base as _leer_base_info, iter_alumnos_hoja, escribir_rangos_consolidado
from .calc_service import promedio_parciales, nota_a_letra

def leer_base(upload_id: str) -> tuple[str, str, list[str]]:
    """
    Retorna (path, filename, sheetnames).
    En productivo: resolver upload_id -> ruta real.
    En S2: se asume upload_id ya es la ruta al .xlsx.
    """
    info = _leer_base_info(upload_id)
    return (str(info.path), info.filename, list(info.sheetnames))

def construir_consolidado(upload_id: str, grado: str, curso: str | None = None) -> List[dict]:
    """
    Lee el archivo y arma el consolidado filtrando por hoja 'grado'
    (si no existe, usa la primera). Filtra por 'curso' si está en hoja.
    Retorna: [{alumnoId, alumno, curso, B1..B4, promedio, letra}]
    """
    path, _, _ = leer_base(upload_id)
    wb = load_workbook(path, data_only=True)
    sheet = grado if grado and grado in wb.sheetnames else wb.sheetnames[0]
    out: List[dict] = []

    for row in iter_alumnos_hoja(wb, sheet):
        if curso and row.get("curso"):
            if str(row["curso"]).strip().lower() != str(curso).strip().lower():
                continue
        prom = promedio_parciales(row.get("B1"), row.get("B2"), row.get("B3"), row.get("B4"))
        out.append({
            "alumnoId": row["alumnoId"],
            "alumno": row["alumno"],
            "curso": row.get("curso") or curso,
            "B1": row.get("B1"),
            "B2": row.get("B2"),
            "B3": row.get("B3"),
            "B4": row.get("B4"),
            "promedio": round(prom, 2),
            "letra": nota_a_letra(prom),
        })
    return out

def exportar_excel(upload_id: str, consolidado: list[dict], comentarios: list[dict]) -> bytes:
    """
    Escribe promedio/letra/comentarios sobre el MISMO archivo,
    manteniendo nombre/hojas/formatos. Retorna bytes del .xlsx.
    Para S2, escribe en la primera hoja (dummy). En real: mapear por grado/curso.
    """
    path, _, _ = leer_base(upload_id)
    wb = load_workbook(path)
    if not wb.sheetnames:
        raise ValueError("El archivo no tiene hojas.")
    hoja = wb.sheetnames[0]

    comentarios_map = {c["alumnoId"]: c.get("texto", "") for c in (comentarios or [])}
    rows = []
    for c in consolidado:
        rows.append({
            "alumnoId": c["alumnoId"],
            "promedio": c.get("promedio"),
            "letra": c.get("letra"),
            "comentario": comentarios_map.get(c["alumnoId"], ""),
        })

    escribir_rangos_consolidado(wb, hoja, rows)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
