# backend/apps/libretas/tests/test_ugel.py
import tempfile
import os
from django.test import TestCase, Client
from django.conf import settings
from openpyxl import Workbook


class UgelEndpointsTests(TestCase):
    def setUp(self):
        self.client = Client()
        settings.USE_FAKE_DATA = True

        # Crear un XLSX dummy para pruebas
        self.temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Test")

        ws['A1'] = 'alumnoId'
        ws['B1'] = 'alumno'
        ws['C1'] = 'B1'
        ws['D1'] = 'B2'
        ws['E1'] = 'B3'
        ws['F1'] = 'B4'
        ws['G1'] = 'curso'
        ws['H1'] = 'promedio'
        ws['I1'] = 'letra'
        ws['J1'] = 'comentario'

        ws['A2'] = 1
        ws['B2'] = 'Test Alumno'
        ws['C2'] = 15
        ws['D2'] = 14
        ws['E2'] = 16
        ws['F2'] = 15
        ws['G2'] = 'Curso A'

        wb.save(self.temp_xlsx.name)
        self.temp_xlsx.close()

    def tearDown(self):
        if os.path.exists(self.temp_xlsx.name):
            os.unlink(self.temp_xlsx.name)

    def test_upload_ok_devuelve_token_y_filename(self):
        """POST /libretas/ugel/upload → 200 con token y filename original"""
        with open(self.temp_xlsx.name, 'rb') as f:
            resp = self.client.post('/api/libretas/ugel/upload', {
                'file': f,
            })

        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn('token', data)
        self.assertIn('filename', data)
        self.assertTrue(data['token'])
        self.assertEqual(data['filename'], os.path.basename(self.temp_xlsx.name))

    def test_upload_falta_archivo(self):
        """POST /libretas/ugel/upload sin archivo → 400"""
        resp = self.client.post('/api/libretas/ugel/upload', {})
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertEqual(data['code'], 'FILE_REQUIRED')

    def test_upload_archivo_invalido(self):
        """POST /libretas/ugel/upload con no-.xlsx → 400"""
        txt_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
        txt_file.write(b'test')
        txt_file.close()

        try:
            with open(txt_file.name, 'rb') as f:
                resp = self.client.post('/api/libretas/ugel/upload', {
                    'file': f,
                })
            self.assertEqual(resp.status_code, 400)
            data = resp.json()
            self.assertEqual(data['code'], 'INVALID_FILE_TYPE')
        finally:
            os.unlink(txt_file.name)

    def test_download_ok_devuelve_xlsx(self):
        """GET /libretas/ugel/download?token=... → 200 con XLSX real"""
        # Primero upload
        with open(self.temp_xlsx.name, 'rb') as f:
            upload_resp = self.client.post('/api/libretas/ugel/upload', {'file': f})

        self.assertEqual(upload_resp.status_code, 200)
        token = upload_resp.json()['token']
        original_filename = upload_resp.json()['filename']

        # Luego download
        download_resp = self.client.get(f'/api/libretas/ugel/download?token={token}')
        self.assertEqual(download_resp.status_code, 200)
        self.assertEqual(
            download_resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Verificar que el filename en Content-Disposition es el original
        self.assertIn(original_filename, download_resp['Content-Disposition'])

    def test_download_token_invalido(self):
        """GET /libretas/ugel/download?token=invalid → 400"""
        resp = self.client.get('/api/libretas/ugel/download?token=invalid_token')
        self.assertEqual(resp.status_code, 400)
        data = resp.json()
        self.assertEqual(data['code'], 'TOKEN_INVALID')

    def test_excel_ok_devuelve_xlsx(self):
        """GET /libretas/ugel/excel?grado=..&seccion=..&anio=.. → 200 con XLSX"""
        resp = self.client.get('/api/libretas/ugel/excel?grado=1&seccion=A&anio=2025')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Verificar que hay un filename
        self.assertIn('UGEL_G1_SA_2025.xlsx', resp['Content-Disposition'])

    def test_upload_then_download_preserva_nombre_y_contenido(self):
        """Test completo: upload → download preserva nombre y escribe Prom/Letra correctamente"""
        from io import BytesIO
        from openpyxl import load_workbook

        # Crear XLSX con datos de prueba
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("1")
        ws.title = "1"
        ws.append(["AlumnoID", "Alumno", "B1", "B2", "B3", "B4", "Curso", "Prom", "Letra", "Comentario"])
        ws.append([1, "Ana", 15, 16, None, None, "Mat"])     # Prom 15.5 -> A
        ws.append([2, "Bruno", 18, 19, 17, None, "Mat"])    # Prom 18.0 -> AD

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        original = "RegNotas_09041770_20_A22025_B1_9999.xlsx"

        # Crear SimpleUploadedFile para simular upload
        from django.core.files.uploadedfile import SimpleUploadedFile
        uploaded_file = SimpleUploadedFile(
            original,
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # 1) Upload
        upload_resp = self.client.post('/api/libretas/ugel/upload', {
            'file': uploaded_file,
        })
        self.assertEqual(upload_resp.status_code, 200)
        data = upload_resp.json()
        token = data["token"]
        self.assertIn(original, data["filename"])

        # 2) Download
        download_resp = self.client.get(f'/api/libretas/ugel/download?token={token}')
        self.assertEqual(download_resp.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                      download_resp["Content-Type"])
        self.assertIn(original, download_resp["Content-Disposition"])

        # 3) Validar que escribió Prom/Letra
        wb_result = load_workbook(filename=BytesIO(download_resp.content), data_only=True)
        ws_result = wb_result["1"]
        
        # Verificar Ana (fila 2)
        prom_ana = ws_result["H2"].value
        letra_ana = ws_result["I2"].value
        self.assertIsNotNone(prom_ana)
        self.assertEqual(letra_ana, "A")
        
        # Verificar Bruno (fila 3)
        prom_bruno = ws_result["H3"].value
        letra_bruno = ws_result["I3"].value
        self.assertIsNotNone(prom_bruno)
        self.assertEqual(letra_bruno, "AD")
