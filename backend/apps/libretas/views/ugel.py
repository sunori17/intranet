# apps/libretas/views/ugel.py
from __future__ import annotations
import json
import os
import tempfile
import uuid
from pathlib import Path
from typing import Optional

from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.conf import settings

from ..services.ugel_service import leer_base, construir_consolidado, exportar_excel
from ..services.pdf_prechecks import verificar_cierre_bimestre


# Almacenamiento temporal para uploads (en real: usar S3 o directorio MEDIA_ROOT)
UPLOAD_TMP_DIR = getattr(settings, 'UPLOAD_TMP_DIR', tempfile.gettempdir())
os.makedirs(UPLOAD_TMP_DIR, exist_ok=True)

# Mapeo token -> (path, original_filename)
_UPLOAD_TOKENS: dict[str, tuple[str, str]] = {}


@require_http_methods(["POST"])
def ugel_upload_post(request):
    """
    POST /libretas/ugel/upload
    Recibe .xlsx (request.FILES["file"]), guarda en UPLOAD_TMP_DIR,
    preserva original_filename, devuelve token y filename.
    Respuesta: {"token": "<uuid>", "filename": "<original>"}
    """
    if "file" not in request.FILES:
        return JsonResponse(
            {"code": "FILE_REQUIRED", "detail": "Falta archivo en request.FILES['file']"},
            status=400
        )

    uploaded_file = request.FILES["file"]
    original_filename = uploaded_file.name

    # Validar extensión
    if not original_filename.lower().endswith(".xlsx"):
        return JsonResponse(
            {"code": "INVALID_FILE_TYPE", "detail": "Solo archivos .xlsx permitidos"},
            status=400
        )

    # Generar token y guardar
    token = str(uuid.uuid4())
    temp_path = os.path.join(UPLOAD_TMP_DIR, f"{token}.xlsx")

    try:
        with open(temp_path, "wb") as f:
            for chunk in uploaded_file.chunks():
                f.write(chunk)
        _UPLOAD_TOKENS[token] = (temp_path, original_filename)
    except Exception as e:
        return JsonResponse(
            {"code": "UPLOAD_ERROR", "detail": str(e)},
            status=500
        )

    return JsonResponse(
        {"token": token, "filename": original_filename},
        status=200
    )


@require_http_methods(["GET"])
def ugel_download_get(request):
    """
    GET /libretas/ugel/download?token=<token>
    Resuelve token -> (path, original_filename),
    llama a consolidar + exportar_excel,
    responde con XLSX real y Content-Disposition preservando nombre original.
    """
    token = request.GET.get("token", "").strip()
    if not token or token not in _UPLOAD_TOKENS:
        return JsonResponse(
            {"code": "TOKEN_INVALID", "detail": "Token no encontrado o expirado"},
            status=400
        )

    path, original_filename = _UPLOAD_TOKENS[token]

    # Opcional: aplicar validaciones de precondición
    # if not settings.USE_FAKE_DATA:
    #     if not verificar_cierre_bimestre(...):
    #         return JsonResponse({...}, status=400)

    try:
        # Construir consolidado desde el archivo
        consolidado = construir_consolidado(path, grado="1", curso="")
        # Exportar XLSX con resultados
        xlsx_bytes = exportar_excel(path, consolidado, comentarios=[])

        resp = HttpResponse(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{original_filename}"'
        return resp
    except Exception as e:
        return JsonResponse(
            {"code": "PROCESSING_ERROR", "detail": str(e)},
            status=500
        )


@require_http_methods(["GET"])
def ugel_excel_get(request):
    """
    GET /libretas/ugel/excel?grado=..&seccion=..&anio=..
    (Opcional) Genera un XLSX directamente sin upload,
    para demo o descarga manual de consolidados.
    Por ahora: responde con XLSX fake o se salta.
    """
    grado = request.GET.get("grado", "1").strip()
    seccion = request.GET.get("seccion", "A").strip()
    anio = request.GET.get("anio", "2025").strip()

    # En real, construirías consolidado desde BD
    # Para S2: mock simple
    try:
        # Mock: crear un XLSX vacío o mínimo
        from openpyxl import Workbook
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Grado 1")
        ws.title = f"Grado {grado}"
        ws["A1"] = "alumnoId"
        ws["B1"] = "alumno"
        ws["C1"] = "B1"
        ws["D1"] = "B2"
        ws["E1"] = "B3"
        ws["F1"] = "B4"
        ws["G1"] = "curso"
        ws["H1"] = "promedio"
        ws["I1"] = "letra"
        ws["J1"] = "comentario"

        stream = BytesIO()
        wb.save(stream)
        xlsx_bytes = stream.getvalue()

        resp = HttpResponse(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"UGEL_G{grado}_S{seccion}_{anio}.xlsx"
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
    except Exception as e:
        return JsonResponse(
            {"code": "GENERATION_ERROR", "detail": str(e)},
            status=500
        )


# Alias para compatibilidad con urls.py (si lo necesitas)
ugel_upload = ugel_upload_post
ugel_download = ugel_download_get
ugel_excel = ugel_excel_get


@require_http_methods(["GET"])
def ugel_consolidado_get(request):
    """
    Mock GET: devuelve un consolidado básico con 2 decimales y letra.
    En real, leerías de BD / excel_adapter.
    """
    if not settings.USE_FAKE_DATA:
        # Si quisieras bloquear en real, aquí pondrías validaciones
        pass

    data = {
        "consolidado": [
            {
                "alumno": "Perez Huaman, Ana Sofía",
                "grado": "1",
                "promedio": 15.25,
                "letra": "A",
            },
            {
                "alumno": "Quispe Lazo, Bruno",
                "grado": "1",
                "promedio": 18.00,
                "letra": "AD",
            },
        ]
    }
    return JsonResponse(data, status=200)


@require_http_methods(["POST"])
def ugel_export_post(request):
    """
    POST /libretas/ugel/export
    (Compatibilidad) Export directo tomando un archivo ya subido (uploadId=token)
    y respondiendo el XLSX con Prom/Letra/Comentario usando servicio real.
    Acepta uploadId, grado, curso por form o JSON.
    """
    payload = {}
    if request.content_type and "application/json" in request.content_type:
        try:
            payload = json.loads(request.body.decode("utf-8"))
        except Exception:
            payload = {}
    upload_id = request.POST.get("uploadId") or payload.get("uploadId")
    
    if not upload_id:
        return JsonResponse(
            {"code": "UPLOAD_ID_REQUIRED", "detail": "Falta uploadId (token del archivo)"},
            status=400
        )

    grado = request.POST.get("grado", "") or payload.get("grado", "")
    curso = request.POST.get("curso", "") or payload.get("curso", "")

    # Si upload_id es un token UUID, resolver desde _UPLOAD_TOKENS
    # Si es una ruta directa, usar directamente
    if upload_id in _UPLOAD_TOKENS:
        path, original_filename = _UPLOAD_TOKENS[upload_id]
    else:
        # Asumir que upload_id es una ruta completa (para compatibilidad)
        path = upload_id
        from ..services.storage import resolve_original_filename
        original_filename = resolve_original_filename(upload_id)

    try:
        # Usar servicio real: construir_consolidado + exportar_excel
        consolidado = construir_consolidado(path, grado=grado, curso=curso)
        comentarios = []  # Mock: cuando conecten BD real, mapear por alumnoId
        xlsx_bytes = exportar_excel(path, consolidado, comentarios)

        resp = HttpResponse(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{original_filename}"'
        return resp
    except Exception as e:
        return JsonResponse(
            {"code": "EXPORT_ERROR", "detail": str(e)},
            status=500
        )
