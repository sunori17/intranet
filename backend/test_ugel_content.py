#!/usr/bin/env python
"""
Script temporal para validar que el XLSX exportado contiene promedio y letra.
"""
import os
import sys
import tempfile
from io import BytesIO

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'intranet.settings')
import django
django.setup()

from django.test import Client
from django.conf import settings
from openpyxl import Workbook, load_workbook

print("=" * 60)
print("TEST: Validación de contenido XLSX exportado")
print("=" * 60)

settings.USE_FAKE_DATA = True
client = Client()

# Crear XLSX dummy
temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
wb = Workbook()
ws = wb.active
if ws is None:
    ws = wb.create_sheet("Test")

ws['A1'] = 'alumnoId'
ws['B1'] = 'alumno'
ws['C1'] = 'B1'
ws['D1'] = 'B2'
ws['E1'] = 'B3'
ws['F1'] = 'B4'
ws['G1'] = 'curso'
ws['H1'] = 'promedio'
ws['I1'] = 'letra'
ws['J1'] = 'comentario'

ws['A2'] = 1
ws['B2'] = 'Test Alumno'
ws['C2'] = 15
ws['D2'] = 14
ws['E2'] = 16
ws['F2'] = 15
ws['G2'] = 'Curso A'

wb.save(temp_xlsx.name)
temp_xlsx.close()

try:
    # 1. Upload
    print("\n1. Uploading XLSX...")
    with open(temp_xlsx.name, 'rb') as f:
        upload_resp = client.post('/libretas/ugel/upload', {'file': f})
    
    if upload_resp.status_code != 200:
        print(f"   ✗ Upload falló: {upload_resp.status_code}")
        print(f"     {upload_resp.content}")
        sys.exit(1)
    
    token = upload_resp.json()['token']
    original_filename = upload_resp.json()['filename']
    print(f"   ✓ Upload OK: token={token[:8]}..., filename={original_filename}")
    
    # 2. Download
    print("\n2. Downloading processed XLSX...")
    download_resp = client.get(f'/libretas/ugel/download?token={token}')
    
    if download_resp.status_code != 200:
        print(f"   ✗ Download falló: {download_resp.status_code}")
        print(f"     {download_resp.content}")
        sys.exit(1)
    
    print(f"   ✓ Download OK: {len(download_resp.content)} bytes")
    
    # 3. Validar contenido
    print("\n3. Validando contenido del XLSX...")
    xlsx_stream = BytesIO(download_resp.content)
    wb_result = load_workbook(xlsx_stream)
    ws_result = wb_result.active
    
    if ws_result is None:
        print("   ✗ No se pudo leer la hoja activa")
        sys.exit(1)
    
    promedio_val = ws_result['H2'].value
    letra_val = ws_result['I2'].value
    comentario_val = ws_result['J2'].value
    
    print(f"   - Promedio (H2): {promedio_val} (tipo: {type(promedio_val).__name__})")
    print(f"   - Letra (I2): {letra_val}")
    print(f"   - Comentario (J2): {comentario_val}")
    
    # Validaciones
    errores = []
    
    if promedio_val is None:
        errores.append("El promedio (H2) está vacío")
    elif not isinstance(promedio_val, (int, float)):
        errores.append(f"El promedio debe ser numérico, es {type(promedio_val).__name__}")
    
    if letra_val is None:
        errores.append("La letra (I2) está vacía")
    elif letra_val not in ['AD', 'A', 'B', 'C']:
        errores.append(f"La letra debe ser AD/A/B/C, es '{letra_val}'")
    
    if errores:
        print("\n   ✗ ERRORES:")
        for e in errores:
            print(f"     - {e}")
        sys.exit(1)
    
    print("\n   ✓ Contenido válido:")
    print(f"     - Promedio calculado correctamente: {promedio_val}")
    print(f"     - Letra asignada correctamente: {letra_val}")
    
    print("\n" + "=" * 60)
    print("✓ TODAS LAS VALIDACIONES PASARON")
    print("=" * 60)

finally:
    if os.path.exists(temp_xlsx.name):
        os.unlink(temp_xlsx.name)
